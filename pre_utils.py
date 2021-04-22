import json
from PIL import Image
import numpy as np
import cv2 
import os 

import glob 
from minio import Minio
from minio.error import S3Error
import time
from random import randrange
from multiprocessing.pool import ThreadPool
from tqdm import tqdm

import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import torch
import segmentation_models_pytorch as smp
import albumentations as albu
from torch.utils.data import DataLoader
from torch.utils.data import Dataset as BaseDataset

import sys
from time import sleep
import openpyxl

from pre_datasets import Dataset1
from pre_datasets import Dataset2
from pre_datasets import Dataset3

# According to Group

def ListOfTarget (t_list) : 

    if t_list == "g1" :
        target_list = ['Cranial Fossa', 'Symphysis',  'Nasal Bone']
    elif t_list == "g2" :
        target_list = ['Maxilla', 'Pterygomaxillary Fissure', 'Orbit']
    elif t_list == "g3" :
        target_list = ['Mandible']

    return target_list

def ListOfDatasets (t_list) : 

    if t_list == "g1" :
        Dataset  = Dataset1 
    elif t_list == "g2" :
        Dataset  = Dataset2 
    elif t_list == "g3" :
        Dataset = Dataset3
    return Dataset 

def ListOfClass (t_list) : 

    if t_list == "g1" :
        CLASSES = ['Cranial Fossa', 'Symphysis',  'Nasal Bone']
    elif t_list == "g2" :
        CLASSES = ['Maxilla', 'Pterygomaxillary Fissure', 'Orbit']
    elif t_list == "g3" :
        CLASSES = ['Mandible']

    return CLASSES

# Create Folder

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def print_progress(iteration, total, prefix='Progress:', suffix='Complete', decimals=1, bar_length=100):
    
    str_format = "{0:." + str(decimals) + "f}"
    current_progress = iteration / float(total)
    percents = str_format.format(100 * current_progress)
    filled_length = int(round(bar_length * current_progress))
    bar = "#" * filled_length + '-' * (bar_length - filled_length)

    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percents, '%', suffix)),

    if iteration == total:
        sys.stdout.write('\n')

    sys.stdout.flush()

def my_IOU_score (gt_mask, pr_mask) : 
    
    (a,b) = gt_mask.shape

    (c,d) = pr_mask.shape

    if (a,b) == (c,d) :

        # Respective Area

        g_count=0
        for i in range(a):
            for j in range(b):
                if gt_mask[i][j] == 1:
                    g_count += 1

        p_count=0
        for i in range(c):
            for j in range(d):
                if pr_mask[i][j] == 1:
                    p_count += 1

        # Overlapping Area

        o_count=0
        for i in range(a):
            for j in range(b):
                if (gt_mask[i][j] == 1) and (pr_mask[i][j] == 1):
                    o_count += 1

    else : 
        print ("error")

    if o_count != 0 :
        prob = (o_count / (g_count + p_count - o_count))*100
    else :
        prob = 0

    return prob

def CreateExel (Basic, name, sheet1 = None, sheet2 = None):
    
    list_added =['Architecture', 'Encoder', 'Latency', 'Throughput', 'CPU', 'GPU', 'Total_Average', 
                        "Cranial Fossa", 'Variance', 'Max', 'Min', 
                        "Symphysis", 'Variance', 'Max', 'Min', 
                        "Nasal Bone",  'Variance', 'Max', 'Min', 
                        "Maxilla",  'Variance', 'Max', 'Min', 
                        "Pterygomaxillary Fissure",  'Variance', 'Max', 'Min', 
                        "Orbit",  'Variance', 'Max', 'Min', 
                        "Mandible", 'Variance', 'Max', 'Min']
    write_wb = openpyxl.Workbook()
    write_ws1 = write_wb.create_sheet(sheet1)
    write_ws1.append(list_added)
    write_ws2 = write_wb.create_sheet(sheet2)
    write_ws2.append(list_added)
    
    name = name + ".xlsx"
    address = os.path.join(Basic,name)
    write_wb.save(address)
    
def print_inf (*arg) :
    
    print ("The Architecture is ", arg[0]) 
    print ("The Encoder is ", arg[1]) 
    print ("The Latency is ", arg[2], "s" )
    print ("The Throughput is ", arg[3], "s")
    print ("The used memory of CPU is ", arg[4] , "MB")
    print ("The used memory of GPU is ", arg[5], "MB")
    print() 

    print ("In total, the Average of the Accuracy is ", np.mean(arg[6]), "%")
    print()    
    
    for i in range(len(arg[6])) : 

        print ("%s Class" % (i+1))
        print ("The Average of the Accuracy is ", np.mean(arg[6][i]), "%")     
        print ("The Variance of the Accuracy is ", np.var(arg[6][i]), "%")     
        print ("The Max of the Accuracy is ", np.max(arg[6][i]), "%")     
        print ("The Min of the Accuracy is ", np.min(arg[6][i]), "%")
        print ()

def visualize(**images):
    """PLot images in one row."""
    n = len(images)
    plt.figure(figsize=(16, 5))
    for i, (name, image) in enumerate(images.items()):
        plt.subplot(1, n, i + 1)
        plt.xticks([])
        plt.yticks([])
        plt.title(' '.join(name.split('_')).title())
        plt.imshow(image)
    plt.show()